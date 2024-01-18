import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
wb = openpyxl.load_workbook("example.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]

#print(get_column_letter(55))
#print(column_index_from_string("FA"))

for row  in sheet["C1":"C7"][::-1]:
    for cell in row:
        print(cell.value)



"""
#for letter in "ABCDEF":
#    for i in range(1,100):
#        sheet[letter + str(i)] = str(datetime.datetime.now())
for i in range(1,sheet.max_row + 1):
    cell = sheet["B"+ str(i)]
    print(i, cell.coordinate, cell.value)
"""

#wb.save("new.xlsx")